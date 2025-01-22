import pandas as pd
import numpy as np
import openpyxl
from BtspStatistics import BtspStatistics
from Statistics_BothAreas import Statistics_BothAreas
from Statistics_Robustness import Statistics_Robustness
from utils import grow_df


class SummaryReport:
    def __init__(self, data_root, output_root, extra_info_CA1, extra_info_CA3, extra_info):
        # set parameters
        self.extra_info_CA1 = extra_info_CA1
        self.extra_info_CA3 = extra_info_CA3
        self.extra_info = "" if not extra_info else f"_{extra_info}"
        self.data_root = data_root
        self.output_root = f"{output_root}//statistics//"

        params = [data_root, output_root, extra_info_CA1, extra_info_CA3, extra_info]
        self.statobj_bothAreas = Statistics_BothAreas(*params, create_output_folder=False)
        self.statobj_robustness = Statistics_Robustness(*params)

        # declare variables
        self.tests_of_interest = ["mann-whitney u", "wilcoxon"]
        self.features = ["initial shift", "log10(formation gain)"]
        self.tests_CA1 = None
        self.tests_CA3 = None
        self.tests_bothAreas = None
        self.equalization_type = "" # min or avg - equlized to min of cells/pfs or avg
        self.equalizations = None
        self.significant_pvalue_proportions = None

    def load_data_and_run_tests(self):
        # load data -- creates BtspStatistics objects for CA1 and CA3
        self.statobj_bothAreas.load_data()

        # run tests for each area and then together
        statobj_CA1 = self.statobj_bothAreas.CA1_stat
        statobj_CA1.run_tests(statobj_CA1.shift_gain_df, params=self.features, export_results=False)
        self.tests_CA1 = statobj_CA1.tests_df

        statobj_CA3 = self.statobj_bothAreas.CA3_stat
        statobj_CA3.run_tests(statobj_CA3.shift_gain_df, params=self.features, export_results=False)
        self.tests_CA3 = statobj_CA3.tests_df

        self.statobj_bothAreas.run_tests(save_results=False)
        self.tests_bothAreas = self.statobj_bothAreas.tests_df

        # filter for tests of interest
        self.tests_CA1 = self.tests_CA1[self.tests_CA1["test"].isin(self.tests_of_interest)]
        self.tests_CA3 = self.tests_CA3[self.tests_CA3["test"].isin(self.tests_of_interest)]
        self.tests_bothAreas = self.tests_bothAreas[self.tests_bothAreas["test"].isin(self.tests_of_interest)]

    def load_equalizations(self, eq_type):
        self.equalization_type = eq_type
        for analysis_type in ["single area", "both areas"]:
            for what_to_eq in ["animal", "session"]:
                for eq_by_what in ["session", "cells", "pfs"]:
                    if what_to_eq == "session" and eq_by_what == "session":
                        continue  # impossible combo
                    eq_df_path = f"{self.output_root}//Robustness//{analysis_type}_equalize_{what_to_eq}_by_{eq_by_what}_to_{eq_type}.pickle"
                    eq_df = pd.read_pickle(eq_df_path)
                    self.equalizations = grow_df(self.equalizations, eq_df)


    def calc_significant_pvalue_proportions(self):
        sps = None
        for _, row in self.equalizations.iterrows():
            analysis_type, what_to_eq, eq_by_what, eq_type, params, tests_df = row.values
            if analysis_type == "both areas":
                tests_df["area"] = "both areas"
            tests_df["is significant"] = tests_df["p-value"] < 0.05
            tests_df = tests_df[(tests_df["test"] == "mann-whitney u") | (tests_df["test"] == "wilcoxon")]

            cols = ["area", "population", "feature", "test"]
            significant_runs = tests_df[[*cols, "is significant", "p-value"]].groupby([*cols, "is significant"]).count()
            total_runs = tests_df[[*cols, "p-value"]].groupby(cols).count()
            significant_proportions = significant_runs / total_runs
            sp = significant_proportions.reset_index().rename(columns={"p-value": "prop"})
            # sp["analysis type"] = analysis_type
            sp["what to eq"] = what_to_eq
            sp["eq by what"] = eq_by_what

            # filter for where significant
            sp = sp[sp["is significant"] == True]

            sp_cols = ["area", "what to eq", "eq by what", "population", "feature", "prop"]
            sp = sp[sp_cols]
            sps = grow_df(sps, sp)
        self.significant_pvalue_proportions = sps

    def create_report(self):
        wb = openpyxl.load_workbook(f"{self.output_root}/summary_template.xlsx")
        ws = wb.active

        tests_rows = {
            "CA1": [self.tests_CA1, 7],
            "CA3": [self.tests_CA3, 12],
            "both areas": [self.tests_bothAreas, 17]
        }
        pops_rows = {
            "established": 0,
            "newly formed": 1,
            "reliables": 2
        }
        feats_cols = {
            "initial shift": "D",
            "log10(formation gain)": "L"
        }
        equalizations = [
            ["animal", "session"],
            ["animal", "cells"],
            ["animal", "pfs"],
            ["session", "cells"],
            ["session", "pfs"]
        ]
        for test_name, test_tuple in tests_rows.items():
            tests, test_row = test_tuple
            for pop, pop_row in pops_rows.items():
                for feat, feat_col in feats_cols.items():
                    if test_name == "both areas" and pop == "reliables":
                        continue  # does not apply; for CA1 v CA3 analysis only either ES or NF can be compared
                    tests_idxed = tests[["population", "feature", "p-value"]].set_index(keys=["population", "feature"])
                    pval = tests_idxed.loc[pop,feat]["p-value"]
                    pval = np.round(pval,3)
                    cell = f"{feat_col}{test_row+pop_row}"
                    ws[cell] = pval

                    sps_idxed = self.significant_pvalue_proportions.set_index(keys=["area", "population", "feature"])
                    props = sps_idxed.loc[test_name, pop, feat]
                    for i_eq, eq_tuple in enumerate(equalizations):
                        what_to_eq, eq_by_what = eq_tuple
                        prop = props[(props["what to eq"] == what_to_eq) & (props["eq by what"] == eq_by_what)]["prop"]
                        if len(prop) == 0:
                            continue
                        prop = np.round(prop.reset_index(drop=True)[0],2)
                        col = chr(ord(feat_col)+1+i_eq)
                        cell = f"{col}{test_row+pop_row}"
                        ws[cell] = prop

        wb.save(f"{self.output_root}/summary_eq_to_{self.equalization_type}.xlsx")

    def calc_tests_by_animals(self):
        ca1_sg = self.statobj_bothAreas.CA1_stat.shift_gain_df
        ca3_sg = self.statobj_bothAreas.CA3_stat.shift_gain_df

        animals_ca1 = ca1_sg["animal id"].unique()
        animals_ca3 = ca3_sg["animal id"].unique()

        params_CA1 = ["CA1", data_root, output_root, extra_info_CA1]
        params_CA3 = ["CA3", data_root, output_root, extra_info_CA3]
        params_bothAreas = [data_root, output_root, extra_info_CA1, extra_info_CA3, extra_info]

        tests_by_animals = None
        for animal in animals_ca1:
            ca1_sg_animal = ca1_sg[ca1_sg["animal id"] == animal]
            ca1_stat_animal = BtspStatistics(*params_CA1)
            ca1_stat_animal.shift_gain_df = ca1_sg_animal
            ca1_stat_animal.run_tests(ca1_stat_animal.shift_gain_df, self.features, export_results=False)
            ca1_tests = ca1_stat_animal.tests_df
            ca1_tests = ca1_tests[ca1_tests["test"].isin(self.tests_of_interest)]
            ca1_tests["animal id"] = animal
            tests_by_animals = grow_df(tests_by_animals, ca1_tests)
        for animal in animals_ca3:
            ca3_sg_animal = ca3_sg[ca3_sg["animal id"] == animal]
            ca3_stat_animal = BtspStatistics(*params_CA3)
            ca3_stat_animal.shift_gain_df = ca3_sg_animal
            ca3_stat_animal.run_tests(ca3_stat_animal.shift_gain_df, self.features, export_results=False)
            ca3_tests = ca3_stat_animal.tests_df
            ca3_tests = ca3_tests[ca3_tests["test"].isin(self.tests_of_interest)]
            ca3_tests["animal id"] = animal
            tests_by_animals = grow_df(tests_by_animals, ca3_tests)
        for animal in np.intersect1d(animals_ca1, animals_ca3):
            stat_bothAreas = Statistics_BothAreas(*params_bothAreas, create_output_folder=False)
            ca1_sg_animal = ca1_sg[ca1_sg["animal id"] == animal]
            ca3_sg_animal = ca3_sg[ca3_sg["animal id"] == animal]
            sg_animal = pd.concat([ca1_sg_animal, ca3_sg_animal])
            stat_bothAreas.shift_gain_df = sg_animal
            stat_bothAreas.run_tests(save_results=False)
            tests_bothAreas = stat_bothAreas.tests_df
            tests_bothAreas = tests_bothAreas[tests_bothAreas["test"].isin(self.tests_of_interest)]
            tests_bothAreas["animal id"] = animal
            tests_by_animals = grow_df(tests_by_animals, tests_bothAreas)
        self.tests_by_animals = tests_by_animals

    def create_report_by_animals(self):
        wb = openpyxl.load_workbook(f"{self.output_root}/tests_animals_template.xlsx")

        area_rows = {
            "CA1": 7,
            "CA3": 12,
            "both": 17
        }
        pops_rows = {
            "established": 0,
            "newly formed": 1,
            "reliables": 2
        }
        feats_cols = {
            "initial shift": "D",
            "log10(formation gain)": "R"
        }
        animals_ordered = ["all", "KS028", "KS029", "KS030", "srb131", "srb231", "srb251", "srb402", "srb410", "srb363", "srb270", "srb269"]
        animal_cols = {
            "initial shift": ["D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"],
            "log10(formation gain)": ["R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC"]
        }
        sheets_columns = {
            "pvalues": "p-value",
            "place cells": "n cells",
            "place fields": "n pfs"
        }

        # merge table of whole data set test results with the one containing test results separately for each animal
        tests_all_animals_CA1 = self.tests_CA1
        tests_all_animals_CA3 = self.tests_CA3
        tests_all_animals_both = self.tests_bothAreas

        tests_all_animals_CA3["animal id"] = "all"
        tests_all_animals_CA1["animal id"] = "all"
        tests_all_animals_both["animal id"] = "all"

        tests_all_animals = pd.concat([tests_all_animals_CA1, tests_all_animals_CA3, tests_all_animals_both])
        tests = pd.concat([self.tests_by_animals, tests_all_animals])
        tests_idxed = tests.set_index(keys=["area", "population", "feature", "animal id"])

        # fill up excel sheets
        for sheet, sheet_df_col in sheets_columns.items():
            ws = wb.get_sheet_by_name(sheet)
            for area, area_row in area_rows.items():
                for pop, pop_row in pops_rows.items():
                    if area == "both" and pop == "reliables":
                        continue  # does not apply; for CA1 v CA3 analysis only either ES or NF can be compared

                    for feat, feat_col in feats_cols.items():
                        for i_animal, animal in enumerate(animals_ordered):
                            animal_idx = animals_ordered.index(animal)
                            try:
                                # treat edge case: CA1 v CA3 and its cell/pf counts (needs to show both areas)
                                if area == "both" and sheet in ["place cells", "place fields"]:
                                    n_ca1 = tests_idxed.loc[area,pop,feat,animal][f"{sheet_df_col} CA1"]
                                    n_ca3 = tests_idxed.loc[area,pop,feat,animal][f"{sheet_df_col} CA3"]
                                    val = f"{int(n_ca1)} / {int(n_ca3)}"
                                else:
                                    df_col = sheet_df_col
                                    if sheet in ["place cells", "place fields"]:
                                        df_col = f"{df_col} {area}"
                                    val = tests_idxed.loc[area, pop, feat, animal][df_col]
                                    val = np.round(val, 3)
                            except KeyError:
                                continue  # thrown for 'illegal' combos, e.g. CA1 + animal from CA3
                            animal_col = animal_cols[feat][animal_idx]
                            cell = f"{animal_col}{area_row+pop_row}"
                            ws[cell] = val
        wb.save(f"{self.output_root}/tests_animals.xlsx")

if __name__ == "__main__":
    data_root = "C:\\Users\\martin\\home\\phd\\btsp_project\\analyses\\manual\\"
    output_root = "C:\\Users\\martin\\home\\phd\\btsp_project\\analyses\\manual\\"

    extra_info_CA1 = ""
    extra_info_CA3 = ""
    extra_info = ""

    report = SummaryReport(data_root, output_root, extra_info_CA1, extra_info_CA3, extra_info)
    report.load_data_and_run_tests()
    report.load_equalizations("avg")
    report.calc_significant_pvalue_proportions()
    report.create_report()

    report.calc_tests_by_animals()
    report.create_report_by_animals()
